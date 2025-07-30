import streamlit as st
import openpyxl
import pandas as pd
import base64
from io import BytesIO

# ========== PAGE CONFIG & CUSTOM STYLE ==========
st.set_page_config(
    page_title="SGPA Calculator",
    page_icon=":bar_chart:",
    layout="centered"
)

st.markdown("""
    <style>
        body {background-color: #161d27;}
        .main {background-color: #161d27 !important;}
        h1, h2, h3, h4, h5, h6, .stMarkdown, .stButton, .stTextInput, .stNumberInput, .stSelectbox, .stFileUploader {
            color: #e0eefd !important;
        }
        .stButton>button {
            background: linear-gradient(90deg, #43cea2, #185a9d);
            border: none;
            color: white;
            font-weight: 600;
            border-radius: 6px;
            padding: 8px 24px;
            margin: 6px 0px;
        }
        .stButton>button:hover {
            background: linear-gradient(90deg, #185a9d, #43cea2);
            color: #fff;
        }
        .stAlert {
            background: #233554;
        }
        .subject-row {
            background: #232d3c;
            border-radius: 6px;
            margin-bottom: 3px;
            padding: 4px 8px;
        }
    </style>
""", unsafe_allow_html=True)

# ========== HEADER ==========
st.markdown(
    """
    <h1 style='text-align: center; color: #43cea2; margin-bottom: 0.1em; letter-spacing:1px;'>
        SGPA Calculator
    </h1>
    """,
    unsafe_allow_html=True
)

st.info("👨‍👩‍👧‍👦 Welcome! Calculate your SGPA by uploading your subject list (Excel) and entering marks below. The calculation follows the official university scheme.", icon="ℹ️")

# ========== DOWNLOADABLE SAMPLE ==========
def create_sample_xlsx():
    output = BytesIO()
    df = pd.DataFrame({
        "Code": ["CS101", "MA102", "PH103", "MCQ201"],
        "Credit": [4, 3, 3, 2],
        "Type": ["Theory", "Theory", "Theory", "MCQ"]
    })
    df.to_excel(output, index=False)
    output.seek(0)
    return output

with st.expander("📥 Download Sample Excel (for demo)"):
    sample_xlsx = create_sample_xlsx()
    b64 = base64.b64encode(sample_xlsx.read()).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="subjects_sample.xlsx">Download sample_subjects.xlsx</a>'
    st.markdown(href, unsafe_allow_html=True)
    st.markdown("""
    ### Sample Excel Format

    | Code   | Credit | Type   |
    |--------|--------|--------|
    | CS101  | 4      | Theory |
    | MA102  | 3      | Theory |
    | PH103  | 3      | Theory |
    | MCQ201 | 2      | MCQ   |
    """, unsafe_allow_html=True)

# ========== EXPLAIN SGPA ==========
with st.expander("📖 What is SGPA? (click to expand)"):
    st.markdown("""
    - **SGPA (Semester Cumulative Grade Point Average)** reflects your overall performance for the semester.
    - Each subject is assigned a **credit** (weightage) and a **type** ("Theory" or "MCQ").
    - For each subject, your marks are converted to a **grade point (out of 10)**, then weighted by the subject's credits.
    - **Formula:**  
      \\[
      \\text{SGPA} = \\frac{\\sum (\\text{Grade Point} \\times \\text{Credit})}{\\sum \\text{Credits}}
      \\]
    """)

# ========== SUBJECT EXCEL UPLOAD ==========
def load_subjects_from_excel(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    subjects = []
    header = [cell.value for cell in ws[1]]
    col_map = {name: idx for idx, name in enumerate(header)}
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        try:
            code = row[col_map.get("Code")]
            credit = row[col_map.get("Credit")]
            type_ = row[col_map.get("Type")]
        except Exception:
            continue
        if str(type_).strip() not in ("Theory", "MCQ"):
            continue
        subjects.append({'code': str(code), 'credit': float(credit), 'type': str(type_)})
    return subjects

uploaded_file = st.file_uploader(
    "Upload your subjects.xlsx (columns: Code, Credit, Type - only 'Theory' or 'MCQ')",
    type=["xlsx"],
    help="Use the sample above if you do not have your own file."
)

# ========== MARKS INPUT ==========
if uploaded_file:
    try:
        subjects = load_subjects_from_excel(uploaded_file)
        if not subjects:
            st.error("No valid subjects found. Only 'Theory' and 'MCQ' types are allowed.")
        else:
            st.success(f"✅ {len(subjects)} subjects loaded! Enter marks below.")
            
            st.markdown("<hr/>", unsafe_allow_html=True)
            st.markdown("### Enter Marks for Each Subject:")

            # Marks input table
            marks_inputs = []
            st.markdown("""
                <div style='display: flex; font-weight: 600; color: #eaeeff;'>
                    <div style='flex:2;'>Code</div>
                    <div style='flex:1;'>Credit</div>
                    <div style='flex:1;'>Type</div>
                    <div style='flex:1;'>Internal <span title="e.g., sessional, quizzes" style="cursor:help;">🛈</span></div>
                    <div style='flex:1;'>Practical <span title="Lab/Assignment marks (if any)" style="cursor:help;">🛈</span></div>
                    <div style='flex:1;'>End Sem <span title="Final university exam" style="cursor:help;">🛈</span></div>
                </div>
            """, unsafe_allow_html=True)
            for idx, subj in enumerate(subjects):
                with st.container():
                    cols = st.columns([2,1,1,1,1,1])
                    cols[0].markdown(f"<div class='subject-row'>{subj['code']}</div>", unsafe_allow_html=True)
                    cols[1].markdown(f"<div class='subject-row'>{subj['credit']}</div>", unsafe_allow_html=True)
                    cols[2].markdown(f"<div class='subject-row'>{subj['type']}</div>", unsafe_allow_html=True)
                    internal = cols[3].number_input(
                        label="",
                        min_value=0, max_value=50, step=1, key=f"internal_{idx}",
                        placeholder="Internal", help="Sessional/Quiz marks"
                    )
                    practical = cols[4].number_input(
                        label="",
                        min_value=0, max_value=50, step=1, key=f"practical_{idx}",
                        placeholder="Practical", help="Lab/Assignment marks"
                    )
                    endsem = cols[5].number_input(
                        label="",
                        min_value=0, max_value=100, step=1, key=f"endsem_{idx}",
                        placeholder="End Sem", help="University Exam marks"
                    )
                    marks_inputs.append({
                        'internal': internal,
                        'practical': practical,
                        'endsem': endsem
                    })

            # ========== CALCULATION BUTTON ==========
            if st.button("🎓 Calculate SGPA", type="primary"):
                total_credits = 0
                weighted_sum = 0
                errors = []
                results_rows = []
                for idx, subj in enumerate(subjects):
                    marks = marks_inputs[idx]
                    if any(m < 0 for m in marks.values()):
                        errors.append(f"Negative marks at subject {subj['code']}")
                        continue
                    total = marks['internal'] + marks['practical'] + marks['endsem']
                    max_marks = 100  # Fixed as per scheme
                    grade_point = min(10, (total / max_marks) * 10)
                    weighted_sum += grade_point * subj['credit']
                    total_credits += subj['credit']
                    results_rows.append({
                        'Code': subj['code'],
                        'Credit': subj['credit'],
                        'Type': subj['type'],
                        'Internal Marks': marks['internal'],
                        'Practical Marks': marks['practical'],
                        'End Sem Marks': marks['endsem'],
                        'Total Marks': total,
                        'Grade Point': round(grade_point, 2)
                    })
                if errors:
                    st.error("Some marks were invalid: " + "; ".join(errors))
                elif total_credits == 0:
                    st.error("No credits found. Please check your Excel file.")
                else:
                    sgpa = weighted_sum / total_credits if total_credits > 0 else 0

                    # ========== ANIMATED RESULT ==========
                    st.balloons()
                    st.markdown(
                        f"""
                        <div style='
                            background: linear-gradient(90deg, #43cea2 0%, #185a9d 100%);
                            color: #fff;
                            padding: 22px 0;
                            border-radius: 10px;
                            font-size: 2.5em;
                            text-align: center;
                            margin-top: 30px;
                        '>
                            🎉 <b>Your SGPA: {sgpa:.2f}</b> 🎉
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                    st.success(
                        "This is your official Semester CGPA as per the entered marks and credits.",
                        icon="🎖️"
                    )
                    with st.expander("How is this calculated?"):
                        st.markdown("""
                            - Grade Point (per subject) = min(10, (Total Marks / Max Marks) × 10)
                            - Weighted by subject credit
                            - SGPA = ∑(grade_point × credit) / ∑credits
                        """)

                    # ========== DOWNLOAD MARKS & SGPA ==========
                    # Add SGPA as last row
                    results_rows.append({
                        'Code': '',
                        'Credit': '',
                        'Type': '',
                        'Internal Marks': '',
                        'Practical Marks': '',
                        'End Sem Marks': '',
                        'Total Marks': '',
                        'Grade Point': f'SGPA: {sgpa:.2f}'
                    })
                    results_df = pd.DataFrame(results_rows)
                    # Export to Excel
                    output_xlsx = BytesIO()
                    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
                        results_df.to_excel(writer, index=False, sheet_name='Results')
                    output_xlsx.seek(0)
                    st.download_button(
                        label="📥 Download Marks & SGPA (Excel)",
                        data=output_xlsx,
                        file_name="SGPA_results.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    # Also provide CSV option
                    output_csv = results_df.to_csv(index=False).encode()
                    st.download_button(
                        label="📥 Download Marks & SGPA (CSV)",
                        data=output_csv,
                        file_name="SGPA_results.csv",
                        mime="text/csv"
                    )
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
else:
    st.info("Please upload a valid Excel file to begin.")

# ========== FOOTER ==========
st.markdown("""
    <hr style="margin:2em 0;"/>
    <div style="text-align:center; color:#6f88a6; font-size:1em;">
        Made by Mohit Kumar
    </div>
""", unsafe_allow_html=True)
